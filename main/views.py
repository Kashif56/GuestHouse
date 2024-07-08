from django.shortcuts import render, get_object_or_404 , redirect, HttpResponseRedirect
from django.contrib import messages
from django.core.exceptions import ValidationError
from .validators import validate_file_extension
from openpyxl.styles import Font, PatternFill, Border, Side

from openpyxl import Workbook
from django.http import HttpResponse


from .models import PurchaseOrder, Invoice, DeliveryChallan, Quote

# Create your views here.

DEPARTMENTS = {
    'Leaf' : 'Leaf',
    'Finance': 'Finance',
    'Security': 'Security',
    'LAB': 'LAB',
    'SMD' : 'SMD',
    'PMD': 'PMD',
    'Engineering' : 'Engineering',
    'Quality': 'Quality',
}

PO_CREATER = [
    'Amina Sardar' , 
    'Khan Muhammad', 
    'Atta Ullah', 
    'Najam', 
    'Saad Hafeez', 
    'Hassan Amin',
    'Mian Hameed','PTC', 
    'Other',
]



def index(request):
    return render(request, 'Home.html')

# ==============================================================
#                          PurchaseOrder
# ==============================================================
def po_list(request):
    if request.method == 'GET':
        search_query = request.GET.get('search_query')
        search_method = request.GET.get('method')
        results = None
        if search_query and search_method:
            if search_method == 'by_po':
                # Filter by PO Number
                results = PurchaseOrder.objects.filter(po_number__exact=search_query).order_by('-id')
            elif search_method == 'by_dept':
                # Filter by Department
                results = PurchaseOrder.objects.filter(department__iexact=search_query).order_by('-id')
            elif search_method == 'by_desc':
                # Filter by Description
                results = PurchaseOrder.objects.filter(description__icontains=search_query).order_by('-id')
       


    po_list = PurchaseOrder.objects.all().order_by('-id')

    
    context = {
        'po_list':po_list,
        'results': results,
        'search_query': search_query,
        'method': search_method,
    }
    
    return render(request, 'po/po_page.html',context)

def po_detail_page(request, po_number):
    context = {}
    po_item = get_object_or_404(PurchaseOrder, po_number=po_number)
    context['po_item'] = po_item

    DCs = po_item.dc.all()
    invoices = po_item.invoices.all()

    # CHecking If details macth or Not

    
    # If DC Count and INV count is same
    if DCs.count() < invoices.count() or DCs.count() > invoices.count():
                count_error_msg = f"DC Count and Invoice did not match"
                context['count_error_msg'] = count_error_msg
    else:
        for dc, invoice in zip(DCs, invoices):
            if dc.dc_number != invoice.invoice_number:
                refNo_error_msg = f"DC number: {dc.dc_number} and Invoice number: {invoice.invoice_number} do not match"
                context['refNo_error_msg'] = refNo_error_msg
        
            if dc.total_amount != invoice.total_amount:
                amount_error_msg = f"Amount at DC Number:{dc.dc_number} - Amount: Rs. {dc.total_amount} and Amount at Invoice number: {invoice.invoice_number} - Amount: Rs. {invoice.total_amount}  do not match"
                context['amount_error_msg'] = amount_error_msg
            
            if dc.dc_date != invoice.invoice_date or invoice.invoice_date > dc.dc_date:
                date_error_msg = f"Date at DC Number:{dc.dc_number} - Date: Rs. {dc.dc_date.strftime("%d, %B %Y")} and Amount at Invoice number: {invoice.invoice_number} - Date: Rs. {invoice.invoice_date.strftime("%d, %B %Y")}  do not match"
                context['date_error_msg'] = date_error_msg
               
                
                
    return render(request, 'po/po_detail_page.html', context)

def po_form_page(request):
    context = {}
    context['Departments'] = DEPARTMENTS
    context['po_creater'] = PO_CREATER
    if request.method == 'POST':
        po_number = request.POST['po_number']
        description = request.POST['description']
        po_date = request.POST['po_date']
        department = request.POST['department']
        po_amount = request.POST['po_amount']
        po_creater = request.POST['po_creater']
     
        po = PurchaseOrder(
            po_number=po_number,
            description=description,
            po_date=po_date,
            department=department,
            po_amount=po_amount,
            po_creater=po_creater,
        )
        po.save()
        messages.success(request, 'Purchase Order has been added successfully.')
        return redirect('po_form')
    
    return render(request, 'po/po_form.html', context)

def edit_po(request, po_number):
    context = {}
    context['Departments'] = DEPARTMENTS
    po = get_object_or_404(PurchaseOrder, po_number=po_number)
    pos = PurchaseOrder.objects.all()
    context['item'] = po
    context['items'] = pos

    if request.method == 'POST':
        po_number = request.POST['po_number']
        description = request.POST['description']
        po_date = request.POST['po_date']
        department = request.POST['department']
        po_amount = request.POST['po_amount']
        po_creater = request.POST['po_creater']
      
        po.po_creater = po_creater
        po.po_number = po_number
        po.description = description
        po.po_date = po_date
        po.department = department
        po.po_amount = po_amount
        po.save()
        
 
        messages.success(request, 'Purchase Order has changed successfully.')
        
        return redirect('po_detail_page', po.po_number)
    return render(request, 'po/po_form.html', context)

def delete_po(request, po_number):
    po = get_object_or_404(PurchaseOrder, po_number=po_number)
    po.delete()

    messages.success(request, "Purchase Order has been deleted Successfully.")

    return redirect('po_list')

# ==============================================================
#                          Quotes
# ==============================================================

def quotes_list(request):
    search_query = request.GET.get('search_query')
    search_method = request.GET.get('method')
    results = None  # Initialize results as None

    if search_query and search_method:
        if search_method == 'by_quote':
            # Filter by Quote Number
            results = Quote.objects.filter(quote_number__exact=search_query).order_by('-id')
        elif search_method == 'by_dept':
            # Filter by Department
            results = Quote.objects.filter(department__iexact=search_query).order_by('-id')
        elif search_method == 'by_desc':
            # Filter by Description
            results = Quote.objects.filter(description__icontains=search_query).order_by('-id')

    # Fetch all Quote objects
    list = Quote.objects.all().order_by('-id')

    context = {
        'items': list,
        'results': results,
        'search_query': search_query,
        'method': search_method,
    }

    return render(request, 'quote/quotes_page.html', context)

def quote_detail_page(request, quote_number):
    item = get_object_or_404(Quote, quote_number=quote_number)
    return render(request, 'quote/quote_detail_page.html', {'item': item})

def quote_form_page(request):
    pos = PurchaseOrder.objects.all()
    context = {}
    context['Departments'] = DEPARTMENTS


    if request.method == 'POST':
        quote_number = request.POST['quote_number']
        quote_img = request.FILES.get('quote_img')
        description = request.POST['description']
        quote_date = request.POST['quote_date']
        event_date = request.POST['event_date']
        department = request.POST['department']
        total_amount = request.POST['total_amount']
        is_po_created = request.POST.get('is_po_created') == 'on'
        po = request.POST['po']
        if po:
            po = PurchaseOrder.objects.get(po_number=po)
        
        try:
            validate_file_extension(quote_img)
        
            quote = Quote(
                    quote_number=quote_number,
                    quote_img = quote_img,
                    description=description,
                    quote_date=quote_date,
                    event_date=event_date,
                    department=department,
                    total_amount=total_amount,
                    is_po_approved=is_po_created,
                    po=po,
            )

            quote.save()   
            messages.success(request, 'Quotation has been added successfully.')
            
            return redirect('quote_form')
        
        except ValidationError as e:
            messages.error(request, 'File Type not Supported')
            return redirect('quote_form')
    
    context['items'] = pos
    return render(request, 'quote/quote_form.html', context)

def edit_quote(request, quote_number):
    context = {}
    context['Departments'] = DEPARTMENTS
    quote = get_object_or_404(Quote, quote_number=quote_number)
    pos = PurchaseOrder.objects.all()
    context['item'] = quote
    context['items'] = pos

    if request.method == 'POST':
        quote_number = request.POST['quote_number']
        description = request.POST['description']
        quote_date = request.POST['quote_date']
        event_date = request.POST['event_date']
        department = request.POST['department']
        total_amount = request.POST['total_amount']
        is_po_created = request.POST.get('is_po_created') == 'on'
        po = request.POST['po']
        if po:
            po = PurchaseOrder.objects.get(po_number=po)
            quote.po = po
        
       
        quote.quote_number = quote_number
        quote.description = description
        quote.quote_date = quote_date
        quote.event_date = event_date
        quote.department = department
        quote.total_amount = total_amount
        quote.is_po_approved = is_po_created
            

        quote.save()
        
 
        messages.success(request, 'Quotation has changed successfully.')
        
        return redirect('quote_detail_page', quote.quote_number)

def delete_quote(request, quote_number):
    quote = get_object_or_404(Quote, quote_number=quote_number)
    quote.delete()

    messages.success(request, "Quotation has been deleted Successfully.")

    return redirect('quote_list')

# ==============================================================
#                          DeliveryChallan
# ==============================================================

def dc_list(request):
    context = {}
    if request.method == 'GET':
        
        search_query = request.GET.get('search_query')
        search_method = request.GET.get('method')

        # Initialize context with default values
        context = {
            'search_query': search_query,
            'method': search_method,
        }
        
        if search_query and search_method:
            if search_method == 'by_dc':
                # Filter by DC Number
                results = DeliveryChallan.objects.filter(dc_number__exact=search_query).order_by('-id')
            elif search_method == 'by_dept':
                # Filter by Department
                results = DeliveryChallan.objects.filter(department__iexact=search_query).order_by('-id')
            elif search_method == 'by_desc':
                # Filter by Description
                results = DeliveryChallan.objects.filter(description__icontains=search_query).order_by('-id')
            else:
                results = []

            # Add results to context if they exist
            context['results'] = results
       


    list = DeliveryChallan.objects.all().order_by('-id')

    grn_dc_total = 0
    nonGrn_dc_total = 0
    
    # Adding Total Amount based on Condition(IF DC is GRN or Not)
    # Calculate total amounts
    grn_dc_total = sum(item.total_amount for item in list if item.is_grn)
    nonGrn_dc_total = sum(item.total_amount for item in list if not item.is_grn)

  
    context.update({
        'items': list,
        'nonGrn_dc_total': nonGrn_dc_total,
        'grn_dc_total': grn_dc_total,
    })
    
    return render(request, 'dc/dc_page.html', context)

def dc_detail_page(request, dc_number):
    item = DeliveryChallan.objects.get(dc_number=dc_number)
    return render(request, 'dc/dc_detail_page.html', {'item': item})
    

def dc_form_page(request):
    pos = PurchaseOrder.objects.all()
    context = {}
    context['Departments'] = DEPARTMENTS

    if request.method == 'POST':
        dc_number = request.POST['dc_number']
        image = request.FILES.get('image')
        description = request.POST['description']
        dc_date = request.POST['dc_date']
        department = request.POST['department']
        total_amount = request.POST['total_amount']
        is_grn = request.POST.get('is_grn') == 'on'
        grn_id = request.POST['grn_id']
        po_number = request.POST['po']
        
        po = get_object_or_404(PurchaseOrder, po_number=po_number)
        
      
            
        dc = DeliveryChallan(
            dc_number=dc_number,
            image=image,
            description=description,
            dc_date=dc_date,
            department=department,
            total_amount=total_amount,
            is_grn=is_grn,
            grn_id=grn_id,
            po=po,
        )

        dc.save()   

        dcs = DeliveryChallan.objects.filter(po__po_number=po)
        total = 0
        for x in dcs:
            total += x.total_amount

        dc_po_remaining_amount = po.po_amount - total
        dc_grn_amount = po.po_amount - dc_po_remaining_amount

        po.grn_amount_by_dc = dc_grn_amount
        po.po_remaining_amount_by_dc = dc_po_remaining_amount
       

        po.dc.add(dc)
        po.save()
        messages.success(request, 'Delivery Challan has been added successfully.')
            
        return redirect('dc_form')
       
    
    context['items'] = pos
    
    return render(request, 'dc/dc_form.html', context)

def edit_dc(request, dc_number):
    context = {}
    context['Departments'] = DEPARTMENTS
    dc = get_object_or_404(DeliveryChallan, dc_number=dc_number)
    pos = PurchaseOrder.objects.all()
    context['item'] = dc
    context['items'] = pos

    if request.method == 'POST':
        dc_number = request.POST['dc_number']
        image = request.FILES.get('image')
        description = request.POST['description']
        dc_date = request.POST['dc_date']
        department = request.POST['department']
        total_amount = int(request.POST['total_amount'])
        is_grn = request.POST.get('is_grn') == 'on'
        grn_id = request.POST['grn_id']
        po_number = request.POST['po']
      
        if po_number:
            po = PurchaseOrder.objects.get(po_number=po_number)
            dc.po = po

        if not image:
            image = dc.image
        
        if dc.total_amount != total_amount:
            po.grn_amount_by_dc -= dc.total_amount
            po.grn_amount_by_dc += total_amount

            po.po_remaining_amount_by_dc += dc.total_amount 
            po.po_remaining_amount_by_dc -= total_amount 

            po.save()
      
        dc.dc_number = dc_number
        dc.description = description
        dc.dc_date = dc_date
        dc.image = image
        dc.is_grn = is_grn
        dc.grn_id = grn_id
        dc.department = department
        dc.total_amount = total_amount
        dc.save()
            
        messages.success(request, 'Delivery Challan has changed successfully.')
        return redirect('dc_detail_page', dc.dc_number)

      


    return render(request, 'dc/dc_form.html', context)

def delete_dc(request, dc_number):
    dc = get_object_or_404(DeliveryChallan, dc_number=dc_number)

    po = dc.po
    
    po.invoices.remove(dc)
    po.po_remaining_amount_by_dc += po.total_amount
    po.grn_amount_by_dc = po.po_amount - po.po_remaining_amount_by_dc
   
    po.save()


    dc.delete()

    messages.success(request, "Delivery Challan has been deleted Successfully.")

  
    return redirect('dc_list')





def invoice_list(request):
    context = {}
    
    if request.method == 'GET':
        search_query = request.GET.get('search_query')
        search_method = request.GET.get('method')
        
        if search_query and search_method:
            context['search_query'] = search_query
            context['method'] = search_method

            if search_method == 'by_inv':
                # Filter by Invoice Number
                results = Invoice.objects.filter(invoice_number__exact=search_query).order_by('-id')
            elif search_method == 'by_dept':
                # Filter by Department
                results = Invoice.objects.filter(department__iexact=search_query).order_by('-id')
            elif search_method == 'by_desc':
                # Filter by Description
                results = Invoice.objects.filter(description__icontains=search_query).order_by('-id')
            elif search_method == 'by_po':
                # Filter by PO Number
                results = Invoice.objects.filter(po__po_number__exact=search_query).order_by('-id')
            else:
                results = Invoice.objects.none()  # Handle default case when search_method doesn't match
            
            context['results'] = results
    
    # Fetch all Invoice objects
    items = Invoice.objects.all().order_by('-id')

    # Calculate totals
    received_inv_total = sum(item.total_amount for item in items if item.is_amount_received)
    pending_inv_total = sum(item.total_amount for item in items if not item.is_amount_received)

    # Update context with items and totals
    context.update({
        'items': items,
        'pending_inv_total': pending_inv_total,
        'received_inv_total': received_inv_total,
    })

    return render(request, 'inv/invoices_page.html', context)


def invoice_detail_page(request, invoice_number):
    item = get_object_or_404(Invoice, invoice_number=invoice_number)
    return render(request, 'inv/invoice_detail_page.html', {'item': item})


def invoice_form_page(request):
    pos = PurchaseOrder.objects.all()
    context = {}
    context['Departments'] = DEPARTMENTS

    if request.method == 'POST':
        invoice_number = request.POST['invoice_number']
        image = request.FILES.get('image')
        description = request.POST['description']
        invoice_date = request.POST['invoice_date']
        sent_date = request.POST['sent_date']
        department = request.POST['department']
        total_amount = request.POST['total_amount']
        is_sent = request.POST.get('is_sent') == 'on'
        grn_id = request.POST['grn_id']
        is_amount_received = request.POST.get('is_amount_received') == 'on'
        po_number = request.POST['po']
        
        po = get_object_or_404(PurchaseOrder, po_number=po_number)
        
      
        invoice = Invoice(
            invoice_number=invoice_number,
            image=image,
            description=description,
            invoice_date=invoice_date,
            sent_date=sent_date,
            department=department,
            total_amount=total_amount,
            is_sent=is_sent,
            is_amount_received=is_amount_received,
            po=po,
            grn_id=grn_id,
        )

        invoice.save()   

        delivery_challan = DeliveryChallan.objects.filter(po__po_number=po, total_amount=total_amount, dc_number=invoice_number)

        invoices = Invoice.objects.filter(po__po_number=po)
        total = 0
        for x in invoices:
            total += x.total_amount

            
        po.invoices.add(invoice)
        invoice.dc.add(delivery_challan)
        po_remaining_amount = po.po_amount - total
        grn_amount = po.po_amount - po_remaining_amount
        po.po_remaining_amount = po_remaining_amount
        po.grn_amount = grn_amount
        po.save()
        messages.success(request, 'Invoice has been added successfully.')
        
       
        
        return redirect('invoice_form')
    
    context['items'] = pos
    
    return render(request, 'inv/invoice_form.html', context)

def edit_invoice(request, invoice_number):
    context = {}
    context['Departments'] = DEPARTMENTS
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    pos = PurchaseOrder.objects.all()
    context['item'] = invoice
    context['items'] = pos

    if request.method == 'POST':
        invoice_number = request.POST['invoice_number']
        image = request.FILES.get('image')
        description = request.POST['description']
        invoice_date = request.POST['invoice_date']
        sent_date = request.POST['sent_date']
        department = request.POST['department']
        total_amount = int(request.POST['total_amount'])
        is_sent = request.POST.get('is_sent') == 'on'
        is_amount_received = request.POST.get('is_amount_received') == 'on'
        po_number = request.POST['po']
        grn_id = request.POST['grn_id']
        
        if po_number:
            po = PurchaseOrder.objects.get(po_number=po_number)
            invoice.po = po
        if not image:
            image = invoice.image
        
        if invoice.total_amount != total_amount:
            po.grn_amount -= invoice.total_amount
            po.grn_amount += total_amount

            po.po_remaining_amount += invoice.total_amount 
            po.po_remaining_amount -= total_amount 
        
            po.save()

        invoice.invoice_number = invoice_number
        invoice.description = description
        invoice.invoice_date = invoice_date
        invoice.sent_date = sent_date
        invoice.image = image
        invoice.is_sent = is_sent
        invoice.grn_id = grn_id
        invoice.is_amount_received = is_amount_received
        invoice.department = department
        invoice.total_amount = total_amount
        invoice.save()
            
        messages.success(request, 'Invoice has changed successfully.')
        return redirect('invoice_detail_page', invoice.invoice_number)
        
        
    
    return render(request, 'inv/invoice_form.html', context)

def delete_invoice(request, invoice_number):
    invoice = get_object_or_404(Invoice, invoice_number=invoice_number)
    po = invoice.po
    
    
    po.invoices.remove(invoice)
    po.po_remaining_amount += invoice.total_amount
    po.grn_amount = po.po_amount - po.po_remaining_amount
   
    po.save()
    
    invoice.delete()

    messages.success(request, "Invoice has been deleted Successfully.")

  
    return redirect('invoice_list')





# Extra Features

def grn_dc(request, po_number):
    context = {}
    po = PurchaseOrder.objects.get(po_number=po_number)
    context['po'] = po

    if request.method == 'POST':
        dc_number = request.POST['dc_number']
        image = request.FILES.get('image')
        description = request.POST['description']
        dc_date = request.POST['dc_date']
        grn_id = request.POST['grn_id']
        total_amount = request.POST['total_amount']
        is_grn = request.POST.get('is_grn') == 'on'
 

        if int(total_amount) < po.po_amount:
            dc = DeliveryChallan(
                    dc_number=dc_number,
                    image=image,
                    description=description,
                    dc_date=dc_date,
                    department=po.department,
                    total_amount=total_amount,
                    is_grn=is_grn,
                    po=po,
                    grn_id=grn_id,
            )

            dc.save()   

            dcs = DeliveryChallan.objects.filter(po__po_number=po)
            total = 0
            for x in dcs:
                total += x.total_amount

            dc_po_remaining_amount = po.po_amount - total
            dc_grn_amount = po.po_amount - dc_po_remaining_amount

            po.grn_amount_by_dc = dc_grn_amount
            po.po_remaining_amount_by_dc = dc_po_remaining_amount

            po.dc.add(dc)
            po.save()
            messages.success(request, f'Delivery Challan has been added successfully for PO#{ po_number }.')
            
            return redirect('po_list')

        else:
            messages.error(request, f'DC Amount is higher than PO Amount.')
            return redirect('grn_dc', po.po_number)


    return render(request, 'grn_dc.html', context)


def grn_inv(request, dc_number):
    context = {}
    dc = DeliveryChallan.objects.get(dc_number=dc_number)
    context['dc'] = dc

    if request.method == 'POST':
        invoice_number = request.POST['invoice_number']
        image = request.FILES.get('image')
        description = request.POST['description']
        invoice_date = request.POST['invoice_date']
        sent_date = request.POST['sent_date']
        total_amount = request.POST['total_amount']
        is_sent = request.POST.get('is_sent') == 'on'
        is_amount_received = request.POST.get('is_amount_received') == 'on'
        grn_id = request.POST['grn_id']
     

    
        invoice = Invoice(
            invoice_number=invoice_number,
            image=image,
            description=description,
            invoice_date=invoice_date,
            sent_date=sent_date,
            department=dc.po.department,
            total_amount=total_amount,
            is_sent=is_sent,
            is_amount_received=is_amount_received,
            po=dc.po,
            grn_id=grn_id,
        )

        invoice.save()   


        invoices = Invoice.objects.filter(po__po_number=dc.po.po_number)
        total = 0
        for x in invoices:
            total += x.total_amount

        
        dc.po.invoices.add(invoice)
        po_remaining_amount = dc.po.po_amount - total
        grn_amount = dc.po.po_amount - po_remaining_amount
        dc.po.po_remaining_amount = po_remaining_amount
        dc.po.grn_amount = grn_amount
        dc.po.save()

        dc.po.invoices.add(invoice)
        invoice.dc.add(dc)
        invoice.save()
        dc.po.save()
        messages.success(request, f'Invoice has been added successfully for PO#{ dc.po.po_number }.')
            
        return redirect('po_list')

     


    return render(request, 'grn_inv.html', context)




# Departments View

def department_list(request):


    context = {}
    if request.method == 'GET':
        search_query = request.GET.get('search_query')
        search_method = request.GET.get('method')
        results = {}
        if search_query and search_method:
            context['search_query']= search_query
            context['method']= search_method
           
            if search_method == 'by_dept':
                # Filter by Department
                results = Invoice.objects.filter(department__iexact=search_query).order_by('-id')
                context['results'] = results
           
    
    context['departments'] = DEPARTMENTS
    invoices = Invoice.objects.all().order_by('-id')
    context['invoices'] = invoices

    

    return render(request, 'dept/all_departments_list.html', context)


def export_invoices_to_excel(request):
    # Fetch all Invoice objects
    invoices = Invoice.objects.all().order_by('-id')

    # Create a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="invoices.xlsx"'

    # Create a new Workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Define headers
    headers = [
        'Invoice Number', 'Image', 'Description', 'Invoice Date', 
        'GRN ID', 'Sent Date', 'Department', 'Total Amount', 
        'Is Sent', 'PO', 'Is Amount Received'
    ]

    # Apply styles
    header_font = Font(bold=True, color="FFFFFF")  # White bold font
    header_fill = PatternFill(start_color='000080', end_color='000080', fill_type='solid')  # Navy Blue background
    cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write headers with styles to the first row
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = cell_border

    # Write data rows with styles
    for row_num, invoice in enumerate(invoices, start=2):
        ws.cell(row=row_num, column=1, value=invoice.invoice_number).border = cell_border
        ws.cell(row=row_num, column=2, value=invoice.image.url if invoice.image else '').border = cell_border
        ws.cell(row=row_num, column=3, value=invoice.description).border = cell_border
        ws.cell(row=row_num, column=4, value=invoice.invoice_date.strftime('%Y-%m-%d')).border = cell_border
        ws.cell(row=row_num, column=5, value=invoice.grn_id).border = cell_border
        ws.cell(row=row_num, column=6, value=invoice.sent_date.strftime('%Y-%m-%d') if invoice.sent_date else '').border = cell_border
        ws.cell(row=row_num, column=7, value=invoice.department).border = cell_border
        ws.cell(row=row_num, column=8, value=f'Rs. {invoice.total_amount}').border = cell_border
        ws.cell(row=row_num, column=9, value='Yes' if invoice.is_sent else 'No').border = cell_border
        ws.cell(row=row_num, column=10, value=invoice.po.po_number if invoice.po else '').border = cell_border
        ws.cell(row=row_num, column=11, value='Yes' if invoice.is_amount_received else 'No').border = cell_border

    # Save the workbook
    wb.save(response)

    return response